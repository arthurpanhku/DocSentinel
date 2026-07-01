"""Graphify-inspired local document ingestion for the knowledge base.

The goal is a lightweight, dependency-minimal pipeline:
raw file -> text -> chunks -> entity graph -> Light RAG markdown artifact.
"""

from __future__ import annotations

import csv
import json
import re
import uuid
import zipfile
from dataclasses import dataclass
from html import unescape
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parents[2]
_UPLOAD_ROOT = _ROOT / "knowledge_base" / "uploads"
RAW_DIR = _UPLOAD_ROOT / "raw"
DISTILLED_DIR = _UPLOAD_ROOT / "distilled"
GRAPH_DIR = _UPLOAD_ROOT / "graphs"

_TOKEN_RE = re.compile(r"[A-Za-z][A-Za-z0-9_-]{2,}|[\u4e00-\u9fff]{2,}")
_HEADING_RE = re.compile(r"^\s{0,3}#{1,6}\s+(.+?)\s*$")
_CONTROL_RE = re.compile(r"\b[A-Z]{2,}(?:-[A-Z0-9]+)+\b")


@dataclass(frozen=True)
class ParsedDocument:
    text: str
    parser: str


def ensure_dirs() -> None:
    for folder in (RAW_DIR, DISTILLED_DIR, GRAPH_DIR):
        folder.mkdir(parents=True, exist_ok=True)


def safe_suffix(filename: str) -> str:
    return "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ".txt"


def parse_document(path: Path, suffix: str) -> ParsedDocument:
    suffix = suffix.lower()
    if suffix in {".txt", ".md"}:
        return ParsedDocument(
            path.read_text(encoding="utf-8", errors="replace"), "plain-text"
        )
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        return ParsedDocument(json.dumps(payload, ensure_ascii=False, indent=2), "json")
    if suffix == ".csv":
        return ParsedDocument(_parse_csv(path), "csv")
    if suffix == ".xlsx":
        return ParsedDocument(_parse_xlsx(path), "xlsx")
    if suffix == ".docx":
        return ParsedDocument(_parse_docx(path), "docx")
    if suffix == ".pdf":
        return ParsedDocument(_parse_pdf_light(path), "pdf-light")
    return ParsedDocument(
        path.read_text(encoding="utf-8", errors="replace"), "fallback-text"
    )


def _parse_csv(path: Path) -> str:
    rows: list[str] = []
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle)
        for idx, row in enumerate(reader, start=1):
            rows.append(
                f"Row {idx}: "
                + " | ".join(cell.strip() for cell in row if cell is not None)
            )
    return "\n".join(rows)


def _parse_xlsx(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in workbook.worksheets:
        parts.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell not in (None, "")]
            if values:
                parts.append(" | ".join(values))
    workbook.close()
    return "\n".join(parts)


def _parse_docx(path: Path) -> str:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    return "\n\n".join(paragraphs)


def _parse_pdf_light(path: Path) -> str:
    # This is not a full PDF parser, but it handles many simple PDFs.
    raw = path.read_bytes()
    text = raw.decode("latin-1", errors="ignore")
    candidates = re.findall(r"\(([^()]{3,})\)\s*T[Jj]", text)
    if not candidates:
        candidates = re.findall(r"\(([^()]{3,})\)", text)
    cleaned = [
        unescape(item).replace("\\n", "\n").replace("\\r", "\n").strip()
        for item in candidates
    ]
    joined = "\n".join(item for item in cleaned if item)
    return (
        joined
        or "PDF text extraction produced no readable text. Upload DOCX, TXT, "
        "MD, CSV, or XLSX for richer parsing."
    )


def chunk_text(text: str, max_chars: int = 1200) -> list[str]:
    paragraphs = [part.strip() for part in re.split(r"\n\s*\n", text) if part.strip()]
    chunks: list[str] = []
    current = ""
    for para in paragraphs:
        if len(current) + len(para) + 2 <= max_chars:
            current = f"{current}\n\n{para}".strip()
        else:
            if current:
                chunks.append(current)
            current = para[:max_chars]
    if current:
        chunks.append(current)
    return chunks or [text[:max_chars]]


def build_graph(title: str, chunks: list[str], source: str) -> dict:
    nodes: list[dict] = [
        {"id": "document", "type": "document", "label": title, "source": source}
    ]
    edges: list[dict] = []
    entity_seen: set[str] = set()
    heading_seen: set[str] = set()

    for idx, chunk in enumerate(chunks):
        chunk_id = f"chunk:{idx}"
        nodes.append(
            {
                "id": chunk_id,
                "type": "chunk",
                "label": f"Chunk {idx + 1}",
                "size": len(chunk),
            }
        )
        edges.append({"source": "document", "target": chunk_id, "type": "contains"})

        for heading in _headings(chunk):
            heading_id = f"section:{_slug(heading)}"
            if heading_id not in heading_seen:
                heading_seen.add(heading_id)
                nodes.append({"id": heading_id, "type": "section", "label": heading})
            edges.append(
                {"source": heading_id, "target": chunk_id, "type": "describes"}
            )

        for entity in _entities(chunk):
            entity_id = f"entity:{_slug(entity)}"
            if entity_id not in entity_seen:
                entity_seen.add(entity_id)
                nodes.append({"id": entity_id, "type": "entity", "label": entity})
            edges.append({"source": chunk_id, "target": entity_id, "type": "mentions"})

    return {
        "nodes": nodes,
        "edges": edges,
        "stats": {
            "chunks": len(chunks),
            "entities": len(entity_seen),
            "sections": len(heading_seen),
        },
    }


def _headings(text: str) -> list[str]:
    headings: list[str] = []
    for line in text.splitlines():
        match = _HEADING_RE.match(line)
        if match:
            headings.append(match.group(1).strip())
    return headings


def _entities(text: str, limit: int = 40) -> list[str]:
    controls = _CONTROL_RE.findall(text)
    words = [
        token
        for token in _TOKEN_RE.findall(text)
        if len(token) >= 4
        and token.lower()
        not in {"with", "that", "this", "from", "should", "must", "shall"}
    ]
    ranked: dict[str, int] = {}
    for token in controls + words:
        key = token.strip()
        ranked[key] = ranked.get(key, 0) + (4 if key in controls else 1)
    return [
        item
        for item, _ in sorted(
            ranked.items(), key=lambda pair: (-pair[1], pair[0].lower())
        )[:limit]
    ]


def _slug(value: str) -> str:
    return (
        re.sub(r"[^A-Za-z0-9_-]+", "-", value.strip())[:80].strip("-").lower()
        or uuid.uuid4().hex
    )


def write_artifacts(
    *, doc_id: uuid.UUID, title: str, filename: str, parsed: ParsedDocument
) -> dict:
    ensure_dirs()
    chunks = chunk_text(parsed.text)
    graph = build_graph(title, chunks, filename)
    markdown = _render_markdown(
        title=title, filename=filename, parser=parsed.parser, chunks=chunks, graph=graph
    )

    md_path = DISTILLED_DIR / f"{doc_id}.md"
    graph_path = GRAPH_DIR / f"{doc_id}.json"
    md_path.write_text(markdown, encoding="utf-8")
    graph_path.write_text(
        json.dumps(graph, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # The Light RAG index is process-local and cached; clear it after new documents.
    from app.services import light_rag

    light_rag._build_index.cache_clear()

    return {
        "markdown_path": str(md_path.relative_to(_ROOT)),
        "graph_path": str(graph_path.relative_to(_ROOT)),
        "chunk_count": len(chunks),
        **graph["stats"],
    }


def _render_markdown(
    *, title: str, filename: str, parser: str, chunks: list[str], graph: dict
) -> str:
    lines = [
        f"# {title}",
        "",
        "## Graphify Metadata",
        f"- source_file: `{filename}`",
        f"- parser: `{parser}`",
        f"- chunks: `{len(chunks)}`",
        f"- entities: `{graph['stats']['entities']}`",
        "",
    ]
    for idx, chunk in enumerate(chunks, start=1):
        lines.extend([f"## Chunk {idx}", "", chunk, ""])
    return "\n".join(lines).strip() + "\n"
