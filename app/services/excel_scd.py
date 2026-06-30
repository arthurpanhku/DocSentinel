"""Compatibility spreadsheet import/export for control verification rows."""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

from app.models.governance import (
    EvidenceItem,
    EvidenceType,
    GateSubmission,
    RequirementRow,
    ReviewStatus,
)
from app.services.excel_contracts import (
    SCD_COLUMN_HEADERS,
    SCD_EXTRAS_IMPLEMENTATION,
    SCD_EXTRAS_REVIEW_DATE,
    SCD_EXTRAS_REVIEWER_NAME,
    SCD_SHEET_NAME,
)

_COL_KEYS = list("ABCDEFGHIJKLM")

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Excel display text <-> DB enum
_REVIEW_STATUS_TO_EXCEL: dict[str, str] = {
    "pending_review": "Pending Review",
    "accepted": "Accepted",
    "rejected": "Rejected",
    "needs_clarification": "Needs Clarification",
}
_EXCEL_TO_REVIEW_STATUS: dict[str, str] = {
    v.lower(): k for k, v in _REVIEW_STATUS_TO_EXCEL.items()
}
# Allow direct snake_case in cell
for k in ReviewStatus.__members__:
    _EXCEL_TO_REVIEW_STATUS[k] = k


def _first_text_evidence(evidences: list[EvidenceItem]) -> tuple[str, str]:
    """Evidence description (G) and reference (H) from first text/link item."""
    for ev in evidences:
        if ev.evidence_type == EvidenceType.text and ev.content:
            return ev.content, ""
        if ev.evidence_type == EvidenceType.link and ev.url:
            return ev.content or "", ev.url or ""
        if ev.file_path:
            return ev.content or "", ev.file_path or ""
    return "", ""


def scd_rows_from_loaded_submission(sub: GateSubmission) -> list[dict[str, Any]]:
    """Build A-M row dicts from a loaded GateSubmission."""
    rows = sorted(sub.requirement_rows, key=lambda r: r.created_at)
    return [requirement_row_to_scd_dict(rr, list(rr.evidence_items)) for rr in rows]


def requirement_row_to_scd_dict(
    row: RequirementRow, evidences: list[EvidenceItem]
) -> dict[str, Any]:
    """Map ORM row + evidence to A–M row dict."""
    g_text, h_ref = _first_text_evidence(evidences)
    extras = row.scd_extras if isinstance(row.scd_extras, dict) else {}
    review_label = _REVIEW_STATUS_TO_EXCEL.get(
        row.review_status.value
        if hasattr(row.review_status, "value")
        else str(row.review_status),
        "Pending Review",
    )
    return {
        "A": row.requirement_id or "",
        "B": row.domain or "",
        "C": row.requirement_text or "",
        "D": row.organization_guidance or "",
        "E": row.applicability or "",
        "F": row.risk_level or "",
        "G": g_text,
        "H": h_ref,
        "I": str(extras.get(SCD_EXTRAS_IMPLEMENTATION, "") or ""),
        "J": row.reviewer_notes or "",
        "K": review_label,
        "L": str(extras.get(SCD_EXTRAS_REVIEW_DATE, "") or ""),
        "M": str(extras.get(SCD_EXTRAS_REVIEWER_NAME, "") or ""),
        "_row_id": str(row.id),
    }


def build_scd_workbook_bytes(rows: list[dict[str, Any]]) -> bytes:
    """Write control verification rows to workbook bytes."""
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SCD_SHEET_NAME

    for col_idx, header in enumerate(SCD_COLUMN_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(_COL_KEYS, start=1):
            val = row.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=val).border = _THIN_BORDER

    ws.freeze_panes = "A2"
    column_widths = [14, 18, 60, 40, 18, 10, 30, 25, 18, 25, 16, 14, 18]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(buf)
    return buf.getvalue()


def parse_scd_workbook(
    data: bytes,
    known_rows: dict[str, RequirementRow],
) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    """
    Parse control verification upload.

    known_rows maps requirement_id (str) to RequirementRow.

    Returns (updates list for DB layer, warnings, errors fatal).
    """
    errors: list[str] = []
    warnings: list[str] = []
    updates: list[dict[str, Any]] = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        return [], [], [f"Invalid Excel file: {exc}"]

    if SCD_SHEET_NAME not in wb.sheetnames:
        # Allow first sheet if name differs slightly
        ws = wb[wb.sheetnames[0]]
        warnings.append(f"Using first sheet '{ws.title}' instead of '{SCD_SHEET_NAME}'")
    else:
        ws = wb[SCD_SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], ["Sheet is empty"]

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    if len(header) < len(SCD_COLUMN_HEADERS):
        return [], [], ["Header row has too few columns"]

    for i, expected in enumerate(SCD_COLUMN_HEADERS):
        if i < len(header) and header[i].lower() != expected.lower():
            warnings.append(
                f"Column {i + 1}: expected '{expected}', got '{header[i] or '?'}'"
            )

    for r_idx, row_vals in enumerate(rows[1:], start=2):
        if not row_vals or all(v is None or str(v).strip() == "" for v in row_vals):
            continue
        cells = list(row_vals) + [None] * (len(_COL_KEYS) - len(row_vals))
        req_id = str(cells[0] or "").strip()
        if not req_id:
            warnings.append(f"Row {r_idx}: skipped (no Requirement ID)")
            continue

        db_row = known_rows.get(req_id)
        if db_row is None:
            warnings.append(
                f"Row {r_idx}: Requirement ID '{req_id}' not found in this submission"
            )
            continue

        review_cell = str(cells[10] or "").strip()
        rs_key = _EXCEL_TO_REVIEW_STATUS.get(review_cell.lower())
        if rs_key is None and review_cell:
            # try partial match
            for k, v in _EXCEL_TO_REVIEW_STATUS.items():
                if review_cell.lower() in k or k in review_cell.lower():
                    rs_key = v
                    break
        if rs_key is None:
            rs_key = (
                db_row.review_status.value
                if hasattr(db_row.review_status, "value")
                else str(db_row.review_status)
            )

        extras: dict[str, Any] = (
            dict(db_row.scd_extras) if isinstance(db_row.scd_extras, dict) else {}
        )
        impl = str(cells[8] or "").strip()
        if impl:
            extras[SCD_EXTRAS_IMPLEMENTATION] = impl
        rd = str(cells[11] or "").strip()
        if rd:
            extras[SCD_EXTRAS_REVIEW_DATE] = rd
        rn = str(cells[12] or "").strip()
        if rn:
            extras[SCD_EXTRAS_REVIEWER_NAME] = rn

        ev_desc = str(cells[6] or "").strip()
        ev_ref = str(cells[7] or "").strip()

        updates.append(
            {
                "requirement_row_id": db_row.id,
                "requirement_id": req_id,
                "domain": str(cells[1] or "").strip() or None,
                "requirement_text": str(cells[2] or "").strip()
                or db_row.requirement_text,
                "organization_guidance": str(cells[3] or "").strip() or None,
                "applicability": str(cells[4] or "").strip() or None,
                "risk_level": str(cells[5] or "").strip() or None,
                "reviewer_notes": str(cells[9] or "").strip() or None,
                "review_status": rs_key,
                "scd_extras": extras,
                "evidence_description": ev_desc,
                "evidence_reference": ev_ref,
            }
        )

    wb.close()
    return updates, warnings, errors
