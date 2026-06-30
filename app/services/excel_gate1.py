"""Compatibility spreadsheet template, import, and export for security intake."""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.excel_contracts import (
    GATE1_ISAF_FIELD_KEYS,
    GATE1_QUESTION_KEYS,
    GATE1_SHEET_ISAF,
    GATE1_SHEET_QUESTIONNAIRE,
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def _normalize_cell(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "yes" if val else "no"
    return str(val).strip()


def build_gate1_template_bytes() -> bytes:
    """Empty security intake workbook."""
    wb = openpyxl.Workbook()
    ws_isaf = wb.active
    ws_isaf.title = GATE1_SHEET_ISAF
    ws_isaf.append(["field_key", "value"])
    for cell in ws_isaf[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for key in GATE1_ISAF_FIELD_KEYS:
        ws_isaf.append([key, ""])

    ws_q = wb.create_sheet(GATE1_SHEET_QUESTIONNAIRE)
    ws_q.append(["question_key", "answer"])
    for cell in ws_q[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for key in GATE1_QUESTION_KEYS:
        ws_q.append([key, ""])

    for ws in (ws_isaf, ws_q):
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 60
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_gate1_workbook_bytes(intake: dict[str, Any]) -> bytes:
    """Export filled security intake data."""
    form = intake.get("form_data") or {}
    decisions = intake.get("decision_questions") or {}

    wb = openpyxl.Workbook()
    ws_isaf = wb.active
    ws_isaf.title = GATE1_SHEET_ISAF
    ws_isaf.append(["field_key", "value"])
    for cell in ws_isaf[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for key in GATE1_ISAF_FIELD_KEYS:
        ws_isaf.append([key, _normalize_cell(form.get(key, ""))])

    ws_q = wb.create_sheet(GATE1_SHEET_QUESTIONNAIRE)
    ws_q.append(["question_key", "answer"])
    for cell in ws_q[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for key in GATE1_QUESTION_KEYS:
        ws_q.append([key, _normalize_cell(decisions.get(key, ""))])

    for ws in (ws_isaf, ws_q):
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_gate1_workbook(data: bytes) -> tuple[dict[str, Any], list[str]]:
    """
    Parse uploaded .xlsx.

    Returns (intake dict {form_data, decision_questions}, warnings).
    Raises ValueError with human-readable message on fatal parse errors.
    """
    warnings: list[str] = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Invalid Excel file: {exc}") from exc

    if GATE1_SHEET_ISAF not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{GATE1_SHEET_ISAF}'")
    if GATE1_SHEET_QUESTIONNAIRE not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{GATE1_SHEET_QUESTIONNAIRE}'")

    form_data: dict[str, str] = {}
    ws_isaf = wb[GATE1_SHEET_ISAF]
    rows_isaf = list(ws_isaf.iter_rows(values_only=True))
    if not rows_isaf:
        raise ValueError("Security intake sheet is empty")
    header = [str(c).lower() if c is not None else "" for c in rows_isaf[0]]
    if "field_key" not in header or "value" not in header:
        raise ValueError("Security intake sheet must have columns field_key, value")
    fk_i = header.index("field_key")
    val_i = header.index("value")
    seen_keys: set[str] = set()
    for row in rows_isaf[1:]:
        if not row or row[fk_i] is None:
            continue
        k = str(row[fk_i]).strip()
        if not k:
            continue
        v = row[val_i] if val_i < len(row) else None
        form_data[k] = _normalize_cell(v)
        seen_keys.add(k)
    for expected in GATE1_ISAF_FIELD_KEYS:
        if expected not in seen_keys:
            warnings.append(
                "Security intake: missing row for field_key "
                f"'{expected}' (optional fill)"
            )

    decisions: dict[str, str] = {}
    ws_q = wb[GATE1_SHEET_QUESTIONNAIRE]
    rows_q = list(ws_q.iter_rows(values_only=True))
    if not rows_q:
        raise ValueError("SecurityQuestionnaire sheet is empty")
    header_q = [str(c).lower() if c is not None else "" for c in rows_q[0]]
    if "question_key" not in header_q or "answer" not in header_q:
        raise ValueError(
            "SecurityQuestionnaire sheet must have columns question_key, answer"
        )
    qk_i = header_q.index("question_key")
    ans_i = header_q.index("answer")
    for row in rows_q[1:]:
        if not row or row[qk_i] is None:
            continue
        k = str(row[qk_i]).strip()
        if not k:
            continue
        v = row[ans_i] if ans_i < len(row) else None
        decisions[k] = _normalize_cell(v)

    wb.close()
    intake = {"form_data": form_data, "decision_questions": decisions}
    return intake, warnings
